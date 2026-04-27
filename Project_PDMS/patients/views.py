import csv
import logging
import openpyxl
from datetime import datetime, date
from dateutil import parser
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db import models
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.views import LoginView, LogoutView
from django.core.paginator import Paginator
from openpyxl import Workbook
from .models import ECHS, CGHS, CAPF
from .utils import get_panel_model, get_expected_headers, get_searchable_fields

logger = logging.getLogger(__name__)

class CustomLoginView(LoginView):
    template_name = "patients/login.html"

    def form_valid(self, form):
        user = form.get_user()
        login(self.request, user)
        logger.info(f"User {user} logged in successfully.")
        return redirect("patients:home")

    def form_invalid(self, form):
        logger.warning("Login failed: Invalid credentials.")
        return super().form_invalid(form)

class CustomLogoutView(LogoutView):
    next_page = "patients:login"

@login_required
def home(request):
    try:
        echs_count = ECHS.objects.values('card_number').distinct().count()
        cghs_count = CGHS.objects.values('card_number').distinct().count()
        capf_count = CAPF.objects.values('card_number').distinct().count()
    except Exception as e:
        logger.error(f"Error fetching dashboard data: {e}")
        return HttpResponse(f"Error querying database: {e}")

    return render(request, "patients/home.html", {
        "echs_count": echs_count,
        "cghs_count": cghs_count,
        "capf_count": capf_count,
    })

@login_required
def panel_dashboard(request, panel_name):
    model, panel_title = get_panel_model(panel_name)
    if not model:
        messages.error(request, "Invalid panel selected.")
        return redirect("patients:home")

    return render(request, "patients/panel_dashboard.html", {
        "panel_name": panel_name,
        "panel_title": panel_title,
    })

@login_required
def upload_data(request, panel_name):
    model, panel_title = get_panel_model(panel_name)
    if not model:
        messages.error(request, "Invalid panel.")
        return redirect("patients:home")

    if request.method == "POST":
        excel_file = request.FILES.get("file")

        if not excel_file or not excel_file.name.endswith(".xlsx"):
            messages.error(request, "Please upload a valid .xlsx file.")
            return redirect("patients:upload_data", panel_name=panel_name)

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            expected_headers = get_expected_headers(panel_name)
            uploaded_headers = [cell.value for cell in sheet[1]]

            if not set(expected_headers).issubset(set(uploaded_headers)):
                missing_headers = set(expected_headers) - set(uploaded_headers)
                messages.error(request, f"Missing required headers: {', '.join(missing_headers)}")
                return redirect("patients:upload_data", panel_name=panel_name)

            column_indices = [uploaded_headers.index(header) for header in expected_headers]
            success_count, duplicate_count = 0, 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {expected_headers[i]: row[column_indices[i]] for i in range(len(column_indices))}

                for key, value in row_data.items():
                    if "date" in key and isinstance(value, str):
                        try:
                            row_data[key] = parser.parse(value).strftime("%Y-%m-%d")
                        except ValueError:
                            row_data[key] = None  

                filters = {key: row_data[key] for key in get_searchable_fields(panel_name) if row_data[key] is not None}
                existing_entry = model.objects.filter(**filters).exists()

                if not existing_entry:
                    try:
                        model.objects.create(**row_data)
                        success_count += 1
                    except Exception as e:
                        logger.error(f"Error inserting row: {e}")
                        messages.error(request, f"Error inserting row: {e}")
                else:
                    duplicate_count += 1

            messages.success(request, f"{success_count} rows uploaded successfully! {duplicate_count} duplicates were skipped.")
            return redirect("patients:upload_data", panel_name=panel_name)

        except Exception as e:
            logger.error(f"Error during upload: {e}")
            messages.error(request, f"Error during upload: {e}")

    return render(request, "patients/upload.html", {"panel_name": panel_name})

@login_required
def search_patient(request, panel_name):
    model, panel_title = get_panel_model(panel_name)
    if not model:
        messages.error(request, "Invalid panel selected.")
        return redirect("patients:home")

    query = request.GET.get("q", "").strip()
    search_by = request.GET.get("search_by", "card_number")
    selected_columns = request.GET.getlist("columns", [])

    filters = {f"{search_by}__icontains": query} if query else {}

    if query:
        results = model.objects.filter(**filters).order_by("-admission_date")
    else:
        results = model.objects.all().order_by("-admission_date")

    paginator = Paginator(results, 20) if results else None
    page = request.GET.get("page")
    paginated_results = paginator.get_page(page) if paginator else None

    fields = [field for field in model._meta.fields if field.name not in ["id"]]
    date_fields = {"admit_date", "admission_date"}
    visible_fields = selected_columns if selected_columns else [field.name for field in fields]

    return render(request, "patients/search.html", {
        "panel_name": panel_name,
        "panel_title": panel_title,
        "results": paginated_results,
        "query": query,
        "search_by": search_by,
        "fields": fields,
        "date_fields": date_fields,
        "visible_fields": visible_fields,
        "search_fields": get_searchable_fields(panel_name),
    })

@login_required
def download_sample_format(request, panel_name):
    model, panel_title = get_panel_model(panel_name)
    if not model:
        messages.error(request, "Invalid panel.")
        return redirect("patients:home")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{panel_name}_sample_format.xlsx"'    

    wb = Workbook()
    ws = wb.active
    ws.title = f"{panel_title} Sample Data"
    ws.append(get_expected_headers(panel_name))

    wb.save(response)
    return response

@login_required
def export_search_results(request, panel_name):
    model, panel_title = get_panel_model(panel_name)
    if not model:
        messages.error(request, "Invalid panel.")
        return redirect("patients:search_patient", panel_name=panel_name)

    query = request.GET.get("q", "").strip()
    search_by = request.GET.get("search_by", "")
    selected_columns = request.GET.getlist("columns", [])

    if not selected_columns:
        selected_columns = [field.name for field in model._meta.fields if field.name != 'id']

    filters = {f"{search_by}__icontains": query}
    results = model.objects.filter(**filters).order_by("-admission_date")

    if not results.exists():
        messages.warning(request, "No matching records found for export.")
        return redirect("patients:search_patient", panel_name=panel_name)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{panel_name}_search_results.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    headers = ['Sr Number'] + [getattr(model._meta.get_field(column), 'verbose_name', column).title() for column in selected_columns]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, result in enumerate(results, start=2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        for col_num, column in enumerate(selected_columns, start=2):
            field = model._meta.get_field(column)
            value = getattr(result, column)
            if isinstance(field, (models.DateField, models.DateTimeField)):
                value = value.strftime("%d-%b-%y") if value else value
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(response)
    return response
